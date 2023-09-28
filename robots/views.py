from datetime import datetime, timedelta
import json
import openpyxl

from django.shortcuts import render
from django.http import HttpResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt

from .models import Robot
from .validators import validate_even


@csrf_exempt
def api_robots(request):
    '''
    API-endpoint для создания в БД записи о новом роботе
    '''

    if request.method == "POST":
        # Загружаем и проверяем на валидность входящие данные
        data: dict = json.loads(request.body.decode())
        validate = validate_even(data)

        # Если данные корректны - сохраняем их в БД
        if validate == True:
            robot: Robot = Robot(
                serial=f"{data['model']}-{data['version']}",
                model=data['model'],
                version=data['version'],
                created=data['created']
            )
            robot.save()

            return HttpResponse('Запись сохраненна')

        # Если нет - отправляем причину ошибки
        return HttpResponse(validate)
    
    if request.method == 'GET':
        return HttpResponse('Чтобы сохранить в БД новую запись о роботе, отправьте POST-запрос с данными в формате JSON')


def total_week(request):
    '''
    Создает и отправлят Excel-файл со сводкой по суммарным показателям производства роботов за последнюю неделю.
    '''

    if request.method == 'GET':
        robots = Robot.objects.all()

        # Вытаскиваем из БД все записи о роботах за последние 7 дней
        last_week = datetime.now() - timedelta(minutes=60*24*7)
        records_week = robots.filter(created__gte=last_week).order_by('serial')

        wb = openpyxl.Workbook()
        serials_checked = [] # серии роботов, что уже внесены в сводку
        row = {}

        for record in records_week:
            # Пробегаемся по записям и сохраняем их в сводку

            serial = record.serial

            if serial in serials_checked:
                continue

            model = record.model
            version = record.version
            count = records_week.filter(serial=serial).count()

            if model in wb.sheetnames:
                sheet = wb[model]

                sheet[f'A{row[model]}'] = model
                sheet[f'B{row[model]}'] = version
                sheet[f'C{row[model]}'] = count

                row[model] += 1
            else:
                if 'Sheet' in wb.sheetnames:
                    wb['Sheet'].title = model
                    sheet = wb[model]
                else:
                    sheet = wb.create_sheet(model)

                sheet['A1'] = 'Model'
                sheet['B1'] = 'Version'
                sheet['C1'] = 'Total'

                sheet['A2'] = model
                sheet['B2'] = version
                sheet['C2'] = count

                row[model] = 3

            wb.save('total_week.xlsx')
            wb.close()

            serials_checked.append(serial)

        return FileResponse(open('total_week.xlsx', 'rb'))
