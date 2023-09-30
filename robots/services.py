from datetime import datetime, timedelta
import json
import openpyxl

from .models import Robot
from .validators import validate_even


def create_new_robot(request):
    # Загружаем и проверяем на валидность входящие данные
    data: dict = json.loads(request.body.decode())
    validate = validate_even(data)

    # Если данные корректны - сохраняем их в БД
    if validate == True:
        robot: Robot = Robot(
            serial=f"{data['model']}-{data['version']}",
            model=data['model'],
            version=data['version'],
            created=data['created']
        )
        robot.save()

        return 'Запись сохраненна'
    
    return validate

def last_week_records(all_objects):
    robots = all_objects
    # Вытаскиваем из БД все записи за последние 7 дней
    last_week = datetime.now() - timedelta(minutes=60*24*7)
    records = robots.filter(created__gte=last_week).order_by('serial')

    return records

def create_summary_robots(all_objects):
    records = last_week_records(all_objects)
    wb = openpyxl.Workbook()
    serials_checked = [] # серии роботов, что уже внесены в сводку
    row = {}

    for record in records:
        # Пробегаемся по записям и сохраняем их в сводку

        serial = record.serial

        if serial in serials_checked:
            continue

        model = record.model
        version = record.version
        count = records.filter(serial=serial).count()

        if model in wb.sheetnames:
            sheet = wb[model]

            sheet[f'A{row[model]}'] = model
            sheet[f'B{row[model]}'] = version
            sheet[f'C{row[model]}'] = count

            row[model] += 1
        else:
            if 'Sheet' in wb.sheetnames:
                wb['Sheet'].title = model
                sheet = wb[model]
            else:
                sheet = wb.create_sheet(model)

            sheet['A1'] = 'Model'
            sheet['B1'] = 'Version'
            sheet['C1'] = 'Total'

            sheet['A2'] = model
            sheet['B2'] = version
            sheet['C2'] = count

            row[model] = 3

        wb.save('total_week.xlsx')
        wb.close()

        serials_checked.append(serial)
